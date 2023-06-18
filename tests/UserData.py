import openpyxl


class User_Data:
    test_userData = [{"name": "user1", "email": "user1@gmail.com", "gender": "Female"},
     {"name": "user2", "email": "user2@gmail.com", "gender": "Male"}]

    @staticmethod
    def get_data(testCaseName):
        book = openpyxl.load_workbook("C://Users/MY/PycharmProjects/SelfFramework_2/Book1.xlsx")
        sheet = book.active
        dict = {}
        list = []
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testCaseName:
                for j in range(2, sheet.max_column + 1):
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [dict]