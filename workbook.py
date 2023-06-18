import openpyxl
book = openpyxl.load_workbook("C://Users/MY/PycharmProjects/SelfFramework_2/Book1.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)
sheet.cell(row=2, column=2).value = "SHAIK SAJID"
print(sheet.cell(row=2, column=2).value)
print(sheet.max_row)
print(sheet.max_column)
print(sheet['D6'].value)
dict = {}
list = []
for i in range(2, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "TestCase1":
        for j in range(2, sheet.max_column + 1):
            dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(dict)